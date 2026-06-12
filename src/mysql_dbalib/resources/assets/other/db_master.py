#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Экзаменационный CLI-мастер для подготовки БД из Excel-ассетов.

Файл специально сделан автономным: его можно скопировать в любую папку,
запустить рядом с проектом и быстро создать SQLite/MySQL-базу.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable


# ============================================================
# РАЗДЕЛ: константы и простые модели данных
# ============================================================

ROLE_NAMES = ("Администратор", "Менеджер", "Авторизированный клиент")
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
MAX_PRODUCT_IMAGE_SIZE = (300, 200)

PRODUCT_COLUMNS = (
    "Артикул",
    "Наименование товара",
    "Единица измерения",
    "Цена",
    "Поставщик",
    "Производитель",
    "Категория товара",
    "Действующая скидка",
    "Кол-во на складе",
    "Описание товара",
    "Фото",
)
USER_COLUMNS = ("Роль сотрудника", "ФИО", "Логин", "Пароль")
ORDER_COLUMNS = (
    "Номер заказа",
    "Артикул заказа",
    "Дата заказа",
    "Дата доставки",
    "Адрес пункта выдачи",
    "ФИО авторизированного клиента",
    "Код для получения",
    "Статус заказа",
)

REFERENCE_TABLES = {
    "roles",
    "categories",
    "suppliers",
    "manufacturers",
    "units",
    "pickup_points",
    "order_statuses",
}

TABLE_ORDER = (
    "roles",
    "users",
    "categories",
    "suppliers",
    "manufacturers",
    "units",
    "products",
    "pickup_points",
    "order_statuses",
    "orders",
    "order_items",
)

TABLE_COLUMNS = {
    "roles": ("id", "name"),
    "users": ("id", "role_id", "full_name", "login", "password"),
    "categories": ("id", "name"),
    "suppliers": ("id", "name"),
    "manufacturers": ("id", "name"),
    "units": ("id", "name"),
    "products": (
        "id",
        "article",
        "name",
        "unit_id",
        "price",
        "supplier_id",
        "manufacturer_id",
        "category_id",
        "discount",
        "stock_quantity",
        "description",
        "photo_path",
    ),
    "pickup_points": ("id", "address"),
    "order_statuses": ("id", "name"),
    "orders": (
        "id",
        "pickup_point_id",
        "client_user_id",
        "client_name",
        "receive_code",
        "status_id",
        "order_date",
        "delivery_date",
    ),
    "order_items": ("id", "order_id", "product_id", "quantity"),
}


class MasterError(RuntimeError):
    """Понятная ошибка для пользователя без длинного traceback."""


@dataclass
class DbSettings:
    engine: str = "sqlite"
    db_path: Path = Path("exam.db")
    database: str = "exam_db"
    host: str = "localhost"
    port: int = 3306
    user: str = "root"
    password: str = ""


@dataclass
class SourceDataset:
    root: Path
    files: dict[str, Path] = field(default_factory=dict)
    rows: dict[str, list[tuple[Any, ...]]] = field(default_factory=dict)
    duplicates: dict[str, list[Path]] = field(default_factory=dict)
    image_files: dict[str, Path] = field(default_factory=dict)


@dataclass
class ScanReport:
    files: dict[str, Path] = field(default_factory=dict)
    missing_files: list[str] = field(default_factory=list)
    duplicates: dict[str, list[Path]] = field(default_factory=dict)
    row_counts: dict[str, int] = field(default_factory=dict)
    importable_counts: dict[str, int] = field(default_factory=dict)
    missing_columns: dict[str, list[str]] = field(default_factory=dict)
    image_count: int = 0
    missing_photos: list[str] = field(default_factory=list)
    invalid_dates: list[str] = field(default_factory=list)
    unknown_order_articles: list[str] = field(default_factory=list)
    bad_order_items: list[str] = field(default_factory=list)


class SourceWorkspace:
    def __init__(self, root: Path, temp_dir: tempfile.TemporaryDirectory[str] | None = None) -> None:
        self.root = root
        self.temp_dir = temp_dir

    def __enter__(self) -> Path:
        return self.root

    def __exit__(self, _exc_type, _exc_value, _traceback) -> None:
        if self.temp_dir:
            self.temp_dir.cleanup()


# ============================================================
# РАЗДЕЛ: SQL-схемы SQLite и MySQL
# ============================================================

SQLITE_SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS roles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    role_id INTEGER NOT NULL,
    full_name TEXT NOT NULL,
    login TEXT NOT NULL UNIQUE,
    password TEXT NOT NULL,
    FOREIGN KEY (role_id) REFERENCES roles(id)
);

CREATE TABLE IF NOT EXISTS categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS suppliers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS manufacturers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS units (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    article TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    unit_id INTEGER NOT NULL,
    price REAL NOT NULL CHECK (price >= 0),
    supplier_id INTEGER NOT NULL,
    manufacturer_id INTEGER NOT NULL,
    category_id INTEGER NOT NULL,
    discount REAL NOT NULL DEFAULT 0 CHECK (discount >= 0 AND discount <= 100),
    stock_quantity INTEGER NOT NULL DEFAULT 0 CHECK (stock_quantity >= 0),
    description TEXT NOT NULL DEFAULT '',
    photo_path TEXT,
    FOREIGN KEY (unit_id) REFERENCES units(id),
    FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
    FOREIGN KEY (manufacturer_id) REFERENCES manufacturers(id),
    FOREIGN KEY (category_id) REFERENCES categories(id)
);

CREATE TABLE IF NOT EXISTS pickup_points (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    address TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS order_statuses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    pickup_point_id INTEGER NOT NULL,
    client_user_id INTEGER,
    client_name TEXT,
    receive_code TEXT,
    status_id INTEGER NOT NULL,
    order_date TEXT,
    delivery_date TEXT,
    FOREIGN KEY (pickup_point_id) REFERENCES pickup_points(id),
    FOREIGN KEY (client_user_id) REFERENCES users(id),
    FOREIGN KEY (status_id) REFERENCES order_statuses(id)
);

CREATE TABLE IF NOT EXISTS order_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL CHECK (quantity > 0),
    FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
    FOREIGN KEY (product_id) REFERENCES products(id),
    UNIQUE (order_id, product_id)
);

CREATE INDEX IF NOT EXISTS idx_products_article ON products(article);
CREATE INDEX IF NOT EXISTS idx_products_name ON products(name);
CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status_id);
CREATE INDEX IF NOT EXISTS idx_order_items_product ON order_items(product_id);
"""

MYSQL_SCHEMA = """
CREATE TABLE IF NOT EXISTS roles (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(100) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS users (
    id INT AUTO_INCREMENT PRIMARY KEY,
    role_id INT NOT NULL,
    full_name VARCHAR(255) NOT NULL,
    login VARCHAR(255) NOT NULL UNIQUE,
    password VARCHAR(255) NOT NULL,
    CONSTRAINT fk_users_role FOREIGN KEY (role_id) REFERENCES roles(id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS categories (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS suppliers (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS manufacturers (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS units (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(100) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS products (
    id INT AUTO_INCREMENT PRIMARY KEY,
    article VARCHAR(100) NOT NULL UNIQUE,
    name VARCHAR(500) NOT NULL,
    unit_id INT NOT NULL,
    price DECIMAL(10,2) NOT NULL,
    supplier_id INT NOT NULL,
    manufacturer_id INT NOT NULL,
    category_id INT NOT NULL,
    discount DECIMAL(5,2) NOT NULL DEFAULT 0,
    stock_quantity INT NOT NULL DEFAULT 0,
    description TEXT NOT NULL,
    photo_path VARCHAR(500),
    INDEX idx_products_article (article),
    INDEX idx_products_name (name),
    CONSTRAINT fk_products_unit FOREIGN KEY (unit_id) REFERENCES units(id),
    CONSTRAINT fk_products_supplier FOREIGN KEY (supplier_id) REFERENCES suppliers(id),
    CONSTRAINT fk_products_manufacturer FOREIGN KEY (manufacturer_id) REFERENCES manufacturers(id),
    CONSTRAINT fk_products_category FOREIGN KEY (category_id) REFERENCES categories(id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS pickup_points (
    id INT AUTO_INCREMENT PRIMARY KEY,
    address VARCHAR(500) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS order_statuses (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(100) NOT NULL UNIQUE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS orders (
    id INT AUTO_INCREMENT PRIMARY KEY,
    pickup_point_id INT NOT NULL,
    client_user_id INT,
    client_name VARCHAR(255),
    receive_code VARCHAR(100),
    status_id INT NOT NULL,
    order_date DATE,
    delivery_date DATE,
    INDEX idx_orders_status (status_id),
    CONSTRAINT fk_orders_pickup FOREIGN KEY (pickup_point_id) REFERENCES pickup_points(id),
    CONSTRAINT fk_orders_user FOREIGN KEY (client_user_id) REFERENCES users(id),
    CONSTRAINT fk_orders_status FOREIGN KEY (status_id) REFERENCES order_statuses(id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS order_items (
    id INT AUTO_INCREMENT PRIMARY KEY,
    order_id INT NOT NULL,
    product_id INT NOT NULL,
    quantity INT NOT NULL,
    UNIQUE KEY uq_order_product (order_id, product_id),
    INDEX idx_order_items_product (product_id),
    CONSTRAINT fk_items_order FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
    CONSTRAINT fk_items_product FOREIGN KEY (product_id) REFERENCES products(id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
"""

MYSQL_DROP_SCRIPT = """
SET FOREIGN_KEY_CHECKS = 0;
DROP TABLE IF EXISTS order_items;
DROP TABLE IF EXISTS orders;
DROP TABLE IF EXISTS products;
DROP TABLE IF EXISTS users;
DROP TABLE IF EXISTS roles;
DROP TABLE IF EXISTS categories;
DROP TABLE IF EXISTS suppliers;
DROP TABLE IF EXISTS manufacturers;
DROP TABLE IF EXISTS units;
DROP TABLE IF EXISTS pickup_points;
DROP TABLE IF EXISTS order_statuses;
SET FOREIGN_KEY_CHECKS = 1;
"""


# ============================================================
# РАЗДЕЛ: подключение к БД
# ============================================================

class CursorAdapter:
    def __init__(self, cursor: Any) -> None:
        self.cursor = cursor
        self.lastrowid = getattr(cursor, "lastrowid", None)

    def fetchone(self) -> dict[str, Any] | None:
        row = self.cursor.fetchone()
        return self._wrap(row)

    def fetchall(self) -> list[dict[str, Any]]:
        return [self._wrap(row) for row in self.cursor.fetchall()]

    def _wrap(self, row: Any) -> dict[str, Any] | None:
        if row is None:
            return None
        if isinstance(row, dict):
            return row
        if isinstance(row, sqlite3.Row):
            return dict(row)
        columns = [description[0] for description in self.cursor.description]
        return dict(zip(columns, row))


class ExamDatabase:
    def __init__(self, settings: DbSettings) -> None:
        self.settings = settings
        self.connection = self._connect()

    def __enter__(self) -> "ExamDatabase":
        return self

    def __exit__(self, exc_type, _exc_value, _traceback) -> None:
        if exc_type:
            self.connection.rollback()
        else:
            self.connection.commit()
        self.connection.close()

    def execute(self, sql: str, params: Iterable[Any] | None = None) -> CursorAdapter:
        cursor = self._cursor()
        cursor.execute(self._prepare_sql(sql), tuple(params or ()))
        return CursorAdapter(cursor)

    def executescript(self, script: str) -> None:
        if self.settings.engine == "sqlite":
            self.connection.executescript(script)
            return

        cursor = self._cursor()
        for statement in split_sql_script(script):
            cursor.execute(statement)

    def commit(self) -> None:
        self.connection.commit()

    def _connect(self):
        if self.settings.engine == "sqlite":
            self.settings.db_path.parent.mkdir(parents=True, exist_ok=True)
            connection = sqlite3.connect(self.settings.db_path)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA foreign_keys = ON")
            return connection

        if self.settings.engine == "mysql":
            connector = import_mysql_connector()
            return connector.connect(
                host=self.settings.host,
                port=self.settings.port,
                user=self.settings.user,
                password=self.settings.password,
                database=self.settings.database,
                charset="utf8mb4",
                use_unicode=True,
            )

        raise MasterError("Движок БД должен быть sqlite или mysql.")

    def _cursor(self):
        if self.settings.engine == "mysql":
            return self.connection.cursor(dictionary=True)
        return self.connection.cursor()

    def _prepare_sql(self, sql: str) -> str:
        if self.settings.engine == "mysql":
            return sql.replace("?", "%s")
        return sql


def import_mysql_connector():
    try:
        import mysql.connector
    except ModuleNotFoundError as error:
        raise MasterError("Установите пакет: pip install mysql-connector-python") from error
    return mysql.connector


def create_mysql_database_if_needed(settings: DbSettings) -> None:
    connector = import_mysql_connector()
    connection = connector.connect(
        host=settings.host,
        port=settings.port,
        user=settings.user,
        password=settings.password,
        charset="utf8mb4",
        use_unicode=True,
    )
    try:
        cursor = connection.cursor()
        cursor.execute(
            f"CREATE DATABASE IF NOT EXISTS `{settings.database}` "
            "CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
        )
        connection.commit()
    finally:
        connection.close()


def initialize_database(settings: DbSettings, reset: bool = False) -> None:
    if settings.engine == "sqlite" and reset and settings.db_path.exists():
        settings.db_path.unlink()

    if settings.engine == "mysql":
        create_mysql_database_if_needed(settings)

    with ExamDatabase(settings) as db:
        if reset and settings.engine == "mysql":
            db.executescript(MYSQL_DROP_SCRIPT)
        db.executescript(MYSQL_SCHEMA if settings.engine == "mysql" else SQLITE_SCHEMA)


def split_sql_script(script: str) -> list[str]:
    statements: list[str] = []
    current: list[str] = []
    in_string = False
    quote_char = ""

    for char in script:
        if char in {"'", '"'}:
            if not in_string:
                in_string = True
                quote_char = char
            elif quote_char == char:
                in_string = False
        if char == ";" and not in_string:
            statement = "".join(current).strip()
            if statement:
                statements.append(statement)
            current = []
        else:
            current.append(char)

    tail = "".join(current).strip()
    if tail:
        statements.append(tail)
    return statements


def db_settings_from_args(args: argparse.Namespace) -> DbSettings:
    engine = getattr(args, "engine", None) or os.environ.get("DB_ENGINE", "sqlite")
    engine = engine.strip().lower()
    return DbSettings(
        engine=engine,
        db_path=Path(getattr(args, "db", None) or os.environ.get("DB_PATH", "exam.db")),
        database=getattr(args, "database", None) or os.environ.get("DB_NAME", "exam_db"),
        host=getattr(args, "host", None) or os.environ.get("DB_HOST", "localhost"),
        port=int(getattr(args, "port", None) or os.environ.get("DB_PORT", "3306")),
        user=getattr(args, "user", None) or os.environ.get("DB_USER", "root"),
        password=getattr(args, "password", None) or os.environ.get("DB_PASSWORD", ""),
    )


# ============================================================
# РАЗДЕЛ: чтение Excel и поиск ассетов
# ============================================================

def open_source_workspace(source: Path) -> SourceWorkspace:
    if not source.exists():
        raise MasterError(f"Источник не найден: {source}")

    if source.is_dir():
        return SourceWorkspace(source)

    temp_dir = tempfile.TemporaryDirectory(prefix="exam_db_assets_")
    extract_root = Path(temp_dir.name)

    if zipfile.is_zipfile(source):
        with zipfile.ZipFile(source) as archive:
            archive.extractall(extract_root)
        return SourceWorkspace(extract_root, temp_dir)

    # На Windows tar часто умеет распаковывать .rar/.zip, если установлен bsdtar.
    result = subprocess.run(
        ["tar", "-xf", str(source), "-C", str(extract_root)],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        temp_dir.cleanup()
        message = result.stderr.strip() or result.stdout.strip()
        raise MasterError(f"Не удалось распаковать архив через tar: {message}")

    return SourceWorkspace(extract_root, temp_dir)


def load_dataset(root: Path) -> SourceDataset:
    dataset = SourceDataset(root=root)
    dataset.image_files = collect_images(root)

    for workbook_path in sorted(root.rglob("*.xlsx")):
        rows = read_excel_rows(workbook_path)
        kind = classify_workbook(workbook_path, rows)
        if kind is None:
            continue

        if kind in dataset.files:
            dataset.duplicates.setdefault(kind, []).append(workbook_path)
            continue

        dataset.files[kind] = workbook_path
        dataset.rows[kind] = rows

    return dataset


def read_excel_rows(path: Path) -> list[tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise MasterError("Установите пакет: pip install openpyxl") from error

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(clean_text(cell) for cell in row):
                rows.append(tuple(row))
        return rows
    finally:
        workbook.close()


def classify_workbook(path: Path, rows: list[tuple[Any, ...]]) -> str | None:
    name = path.name.lower()
    first_row = [clean_text(cell) for cell in rows[0]] if rows else []
    first_row_lower = [value.lower() for value in first_row]

    if name == "tovar.xlsx":
        return "products"
    if name == "user_import.xlsx":
        return "users"
    if "номер заказа" in first_row_lower:
        return "orders"
    if "пункт" in name or "выдач" in name:
        return "pickup_points"
    if first_row and first_row[0].startswith(("1", "2", "3", "4", "5", "6", "7", "8", "9")):
        return "pickup_points"
    return None


def collect_images(root: Path) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES:
            result.setdefault(path.name.lower(), path)
    return result


def find_image_file(dataset: SourceDataset, photo_name: str) -> Path | None:
    name = clean_text(photo_name)
    if not name:
        return None

    direct = dataset.image_files.get(name.lower())
    if direct:
        return direct

    stem = Path(name).stem.lower()
    for image_path in dataset.image_files.values():
        if image_path.stem.lower() == stem:
            return image_path
    return None


def require_dataset(dataset: SourceDataset) -> None:
    missing = [kind for kind in ("products", "users", "orders", "pickup_points") if kind not in dataset.files]
    if missing:
        names = ", ".join(missing)
        raise MasterError(f"Не найдены обязательные Excel-файлы: {names}")

    missing_columns: dict[str, list[str]] = {}
    for kind, columns in (
        ("products", PRODUCT_COLUMNS),
        ("users", USER_COLUMNS),
        ("orders", ORDER_COLUMNS),
    ):
        rows = dataset.rows[kind]
        headers = header_indexes(rows)
        absent = [column for column in columns if column not in headers]
        if absent:
            missing_columns[kind] = absent

    if missing_columns:
        details = "; ".join(
            f"{kind}: {', '.join(columns)}" for kind, columns in missing_columns.items()
        )
        raise MasterError(f"В Excel не хватает колонок: {details}")


def header_indexes(rows: list[tuple[Any, ...]]) -> dict[str, int]:
    if not rows:
        return {}
    return {clean_text(value): index for index, value in enumerate(rows[0]) if clean_text(value)}


def cell(row: tuple[Any, ...], indexes: dict[str, int], column_name: str) -> Any:
    index = indexes[column_name]
    return row[index] if index < len(row) else None


# ============================================================
# РАЗДЕЛ: scan-отчёт по исходным ассетам
# ============================================================

def build_scan_report(dataset: SourceDataset) -> ScanReport:
    report = ScanReport(
        files=dataset.files.copy(),
        missing_files=[
            kind for kind in ("products", "users", "orders", "pickup_points") if kind not in dataset.files
        ],
        duplicates=dataset.duplicates.copy(),
        image_count=len(dataset.image_files),
    )

    for kind, rows in dataset.rows.items():
        report.row_counts[kind] = len(rows)
        report.importable_counts[kind] = count_importable_rows(kind, rows)

    for kind, columns in (
        ("products", PRODUCT_COLUMNS),
        ("users", USER_COLUMNS),
        ("orders", ORDER_COLUMNS),
    ):
        if kind not in dataset.rows:
            continue
        indexes = header_indexes(dataset.rows[kind])
        absent = [column for column in columns if column not in indexes]
        if absent:
            report.missing_columns[kind] = absent

    product_articles = collect_product_articles(dataset)
    report.missing_photos = find_missing_photos(dataset)
    (
        report.invalid_dates,
        report.unknown_order_articles,
        report.bad_order_items,
    ) = inspect_order_rows(dataset, product_articles)
    return report


def count_importable_rows(kind: str, rows: list[tuple[Any, ...]]) -> int:
    if not rows:
        return 0
    if kind == "pickup_points":
        return sum(1 for row in rows if clean_text(row[0] if row else None))

    indexes = header_indexes(rows)
    if kind == "products" and "Артикул" in indexes:
        return sum(1 for row in rows[1:] if clean_text(cell(row, indexes, "Артикул")))
    if kind == "users" and "Логин" in indexes:
        return sum(1 for row in rows[1:] if clean_text(cell(row, indexes, "Логин")))
    if kind == "orders" and "Номер заказа" in indexes:
        return sum(1 for row in rows[1:] if clean_text(cell(row, indexes, "Номер заказа")))
    return max(0, len(rows) - 1)


def collect_product_articles(dataset: SourceDataset) -> set[str]:
    rows = dataset.rows.get("products", [])
    indexes = header_indexes(rows)
    if "Артикул" not in indexes:
        return set()
    return {
        clean_text(cell(row, indexes, "Артикул"))
        for row in rows[1:]
        if clean_text(cell(row, indexes, "Артикул"))
    }


def find_missing_photos(dataset: SourceDataset) -> list[str]:
    rows = dataset.rows.get("products", [])
    indexes = header_indexes(rows)
    if "Артикул" not in indexes or "Фото" not in indexes:
        return []

    missing = []
    for row in rows[1:]:
        article = clean_text(cell(row, indexes, "Артикул"))
        photo_name = clean_text(cell(row, indexes, "Фото"))
        if article and photo_name and find_image_file(dataset, photo_name) is None:
            missing.append(f"{article}: {photo_name}")
    return missing


def inspect_order_rows(dataset: SourceDataset, product_articles: set[str]) -> tuple[list[str], list[str], list[str]]:
    rows = dataset.rows.get("orders", [])
    indexes = header_indexes(rows)
    if not rows or not indexes:
        return [], [], []

    invalid_dates: list[str] = []
    unknown_articles: list[str] = []
    bad_items: list[str] = []

    for row in rows[1:]:
        order_id = clean_text(cell(row, indexes, "Номер заказа"))
        if not order_id:
            continue

        for column_name in ("Дата заказа", "Дата доставки"):
            raw_date = cell(row, indexes, column_name)
            parsed, warning = parse_excel_date(raw_date, order_id)
            if warning:
                invalid_dates.append(f"Заказ {order_id}, {column_name}: {warning}")
            elif raw_date and parsed is None:
                invalid_dates.append(f"Заказ {order_id}, {column_name}: {raw_date}")

        parsed_items, item_warnings = parse_order_items(cell(row, indexes, "Артикул заказа"))
        bad_items.extend(f"Заказ {order_id}: {warning}" for warning in item_warnings)
        for article, quantity in parsed_items:
            if article not in product_articles:
                unknown_articles.append(f"Заказ {order_id}: {article} x {quantity}")

    return invalid_dates, unknown_articles, bad_items


def print_scan_report(report: ScanReport) -> None:
    print("SCAN REPORT")
    print("Файлы:")
    labels = {
        "products": "товары",
        "users": "пользователи",
        "orders": "заказы",
        "pickup_points": "пункты выдачи",
    }
    for kind, label in labels.items():
        if kind in report.files:
            rows = report.row_counts.get(kind, 0)
            importable = report.importable_counts.get(kind, 0)
            print(f"  OK {label}: {report.files[kind]} | строк: {rows}, к импорту: {importable}")
        else:
            print(f"  НЕТ {label}")

    if report.duplicates:
        print("Дубликаты Excel:")
        for kind, paths in report.duplicates.items():
            print(f"  {kind}: {', '.join(str(path) for path in paths)}")

    if report.missing_columns:
        print("Отсутствующие колонки:")
        for kind, columns in report.missing_columns.items():
            print(f"  {kind}: {', '.join(columns)}")

    print(f"Фото найдено: {report.image_count}")
    print_limited("Фото указано в Excel, но не найдено", report.missing_photos)
    print_limited("Некорректные даты", report.invalid_dates)
    print_limited("Позиции заказа с неизвестным артикулом", report.unknown_order_articles)
    print_limited("Проблемы разбора позиций заказа", report.bad_order_items)

    has_problem = bool(
        report.missing_files
        or report.missing_columns
        or report.missing_photos
        or report.invalid_dates
        or report.unknown_order_articles
        or report.bad_order_items
    )
    print("Итог:", "есть предупреждения" if has_problem else "OK")


def print_limited(title: str, values: list[str], limit: int = 12) -> None:
    if not values:
        return
    print(f"{title}:")
    for value in values[:limit]:
        print(f"  - {value}")
    if len(values) > limit:
        print(f"  ... ещё {len(values) - limit}")


# ============================================================
# РАЗДЕЛ: импорт справочников, пользователей и товаров
# ============================================================

def import_all_data(
    db: ExamDatabase,
    dataset: SourceDataset,
    assets_dir: Path,
    report_messages: list[str],
) -> None:
    require_dataset(dataset)
    insert_default_roles(db)
    import_users(db, dataset.rows["users"])
    import_pickup_points(db, dataset.rows["pickup_points"])
    import_products(db, dataset, assets_dir, report_messages)
    import_orders(db, dataset.rows["orders"], report_messages)
    db.commit()


def insert_default_roles(db: ExamDatabase) -> None:
    for role_name in ROLE_NAMES:
        get_or_create_named(db, "roles", role_name)


def import_users(db: ExamDatabase, rows: list[tuple[Any, ...]]) -> None:
    indexes = header_indexes(rows)
    for row in rows[1:]:
        login = clean_text(cell(row, indexes, "Логин"))
        password = clean_text(cell(row, indexes, "Пароль"))
        full_name = clean_text(cell(row, indexes, "ФИО"))
        role_name = clean_text(cell(row, indexes, "Роль сотрудника"))
        if not login or not password or not full_name or not role_name:
            continue

        role_id = get_or_create_named(db, "roles", role_name)
        existing = db.execute("SELECT id FROM users WHERE login = ?", (login,)).fetchone()
        if existing:
            db.execute(
                "UPDATE users SET role_id = ?, full_name = ?, password = ? WHERE login = ?",
                (role_id, full_name, password, login),
            )
            continue

        db.execute(
            "INSERT INTO users(role_id, full_name, login, password) VALUES (?, ?, ?, ?)",
            (role_id, full_name, login, password),
        )


def import_pickup_points(db: ExamDatabase, rows: list[tuple[Any, ...]]) -> None:
    for row in rows:
        address = clean_text(row[0] if row else None)
        if not address or address.lower() in {"адрес", "адрес пункта выдачи"}:
            continue
        get_or_create_pickup_point(db, address)


def import_products(
    db: ExamDatabase,
    dataset: SourceDataset,
    assets_dir: Path,
    report_messages: list[str],
) -> None:
    rows = dataset.rows["products"]
    indexes = header_indexes(rows)

    for row in rows[1:]:
        article = clean_text(cell(row, indexes, "Артикул"))
        if not article:
            continue

        unit_id = get_or_create_named(db, "units", cell(row, indexes, "Единица измерения"))
        supplier_id = get_or_create_named(db, "suppliers", cell(row, indexes, "Поставщик"))
        manufacturer_id = get_or_create_named(db, "manufacturers", cell(row, indexes, "Производитель"))
        category_id = get_or_create_named(db, "categories", cell(row, indexes, "Категория товара"))
        photo_path = prepare_product_image(
            dataset,
            cell(row, indexes, "Фото"),
            article,
            assets_dir,
            report_messages,
        )

        values = (
            clean_text(cell(row, indexes, "Наименование товара")),
            unit_id,
            to_float(cell(row, indexes, "Цена")),
            supplier_id,
            manufacturer_id,
            category_id,
            to_float(cell(row, indexes, "Действующая скидка")),
            to_int(cell(row, indexes, "Кол-во на складе")),
            clean_text(cell(row, indexes, "Описание товара")),
            photo_path,
        )

        existing = db.execute("SELECT id FROM products WHERE article = ?", (article,)).fetchone()
        if existing:
            db.execute(
                """
                UPDATE products
                SET name = ?, unit_id = ?, price = ?, supplier_id = ?, manufacturer_id = ?,
                    category_id = ?, discount = ?, stock_quantity = ?, description = ?, photo_path = ?
                WHERE article = ?
                """,
                (*values, article),
            )
            continue

        db.execute(
            """
            INSERT INTO products(
                article, name, unit_id, price, supplier_id, manufacturer_id,
                category_id, discount, stock_quantity, description, photo_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (article, *values),
        )


def prepare_product_image(
    dataset: SourceDataset,
    raw_photo_name: Any,
    article: str,
    assets_dir: Path,
    report_messages: list[str],
) -> str | None:
    photo_name = clean_text(raw_photo_name)
    if not photo_name:
        return None

    source = find_image_file(dataset, photo_name)
    if source:
        destination_name = f"{source.stem}{source.suffix.lower()}"
    else:
        source_name = Path(photo_name)
        suffix = source_name.suffix.lower() or ".jpg"
        destination_name = f"{source_name.stem}{suffix}"

    destination_dir = assets_dir / "products"
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / destination_name
    relative_path = f"assets/products/{destination_name}"

    if source is None:
        report_messages.append(f"Товар {article}: фото {photo_name} не найдено, сохранён путь {relative_path}.")
        return relative_path

    if source.name != photo_name:
        report_messages.append(f"Товар {article}: вместо {photo_name} найден файл {source.name}.")

    try:
        from PIL import Image
    except ModuleNotFoundError:
        shutil.copy2(source, destination)
        report_messages.append(
            f"Товар {article}: Pillow не установлен, фото скопировано без уменьшения."
        )
        return relative_path

    try:
        with Image.open(source) as image:
            image.thumbnail(MAX_PRODUCT_IMAGE_SIZE)
            save_image = image.convert("RGB") if destination.suffix.lower() in {".jpg", ".jpeg"} else image
            save_image.save(destination)
    except Exception as error:  # noqa: BLE001 - на экзамене важнее не сорвать импорт из-за одного фото.
        shutil.copy2(source, destination)
        report_messages.append(f"Товар {article}: фото не удалось сжать ({error}), скопировано как есть.")

    return relative_path


# ============================================================
# РАЗДЕЛ: импорт заказов
# ============================================================

def import_orders(db: ExamDatabase, rows: list[tuple[Any, ...]], report_messages: list[str]) -> None:
    indexes = header_indexes(rows)

    for row in rows[1:]:
        raw_order_id = clean_text(cell(row, indexes, "Номер заказа"))
        if not raw_order_id:
            continue

        order_id = to_int(raw_order_id)
        status_id = get_or_create_named(db, "order_statuses", cell(row, indexes, "Статус заказа"))
        pickup_point_id = resolve_pickup_point(db, cell(row, indexes, "Адрес пункта выдачи"))
        client_name = clean_text(cell(row, indexes, "ФИО авторизированного клиента"))
        client_user_id = find_user_by_name(db, client_name)
        order_date, order_date_warning = parse_excel_date(cell(row, indexes, "Дата заказа"), str(order_id))
        delivery_date, delivery_date_warning = parse_excel_date(cell(row, indexes, "Дата доставки"), str(order_id))
        receive_code = clean_text(cell(row, indexes, "Код для получения"))

        for warning in (order_date_warning, delivery_date_warning):
            if warning:
                report_messages.append(f"Заказ {order_id}: {warning}")

        existing = db.execute("SELECT id FROM orders WHERE id = ?", (order_id,)).fetchone()
        if existing:
            db.execute(
                """
                UPDATE orders
                SET pickup_point_id = ?, client_user_id = ?, client_name = ?, receive_code = ?,
                    status_id = ?, order_date = ?, delivery_date = ?
                WHERE id = ?
                """,
                (
                    pickup_point_id,
                    client_user_id,
                    client_name,
                    receive_code,
                    status_id,
                    order_date,
                    delivery_date,
                    order_id,
                ),
            )
            db.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))
        else:
            db.execute(
                """
                INSERT INTO orders(
                    id, pickup_point_id, client_user_id, client_name, receive_code,
                    status_id, order_date, delivery_date
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    order_id,
                    pickup_point_id,
                    client_user_id,
                    client_name,
                    receive_code,
                    status_id,
                    order_date,
                    delivery_date,
                ),
            )

        items, item_warnings = parse_order_items(cell(row, indexes, "Артикул заказа"))
        for warning in item_warnings:
            report_messages.append(f"Заказ {order_id}: {warning}")

        for article, quantity in items:
            product_id = find_product_by_article(db, article)
            if product_id is None:
                report_messages.append(
                    f"Заказ {order_id}: позиция {article} x {quantity} пропущена, товара нет в Tovar.xlsx."
                )
                continue

            db.execute(
                "INSERT INTO order_items(order_id, product_id, quantity) VALUES (?, ?, ?)",
                (order_id, product_id, quantity),
            )


def resolve_pickup_point(db: ExamDatabase, value: Any) -> int:
    text = clean_text(value)
    if not text:
        return get_or_create_pickup_point(db, "Не указан")

    if re.fullmatch(r"\d+", text):
        row = db.execute("SELECT id FROM pickup_points WHERE id = ?", (int(text),)).fetchone()
        if row:
            return int(row["id"])

    return get_or_create_pickup_point(db, text)


def find_user_by_name(db: ExamDatabase, full_name: str) -> int | None:
    if not full_name:
        return None
    row = db.execute("SELECT id FROM users WHERE full_name = ?", (full_name,)).fetchone()
    return int(row["id"]) if row else None


def find_product_by_article(db: ExamDatabase, article: str) -> int | None:
    row = db.execute("SELECT id FROM products WHERE article = ?", (article,)).fetchone()
    return int(row["id"]) if row else None


# ============================================================
# РАЗДЕЛ: общие функции импорта и очистки значений
# ============================================================

def get_or_create_named(db: ExamDatabase, table_name: str, raw_name: Any) -> int:
    if table_name not in REFERENCE_TABLES or table_name == "pickup_points":
        raise MasterError(f"Недопустимый справочник: {table_name}")

    name = clean_text(raw_name) or "Не указано"
    row = db.execute(f"SELECT id FROM {table_name} WHERE name = ?", (name,)).fetchone()
    if row:
        return int(row["id"])

    cursor = db.execute(f"INSERT INTO {table_name}(name) VALUES (?)", (name,))
    return int(cursor.lastrowid)


def get_or_create_pickup_point(db: ExamDatabase, raw_address: Any) -> int:
    address = clean_text(raw_address) or "Не указан"
    row = db.execute("SELECT id FROM pickup_points WHERE address = ?", (address,)).fetchone()
    if row:
        return int(row["id"])

    cursor = db.execute("INSERT INTO pickup_points(address) VALUES (?)", (address,))
    return int(cursor.lastrowid)


def parse_order_items(value: Any) -> tuple[list[tuple[str, int]], list[str]]:
    text = clean_text(value)
    if not text:
        return [], ["пустая строка позиций заказа"]

    parts = [part.strip() for part in re.split(r"[,;\n]+", text) if part.strip()]
    items: list[tuple[str, int]] = []
    warnings: list[str] = []

    if len(parts) % 2 != 0:
        warnings.append(f"нечётное число элементов в строке позиций: {text}")

    for index in range(0, len(parts), 2):
        if index + 1 >= len(parts):
            break

        article = clean_text(parts[index])
        quantity = to_int(parts[index + 1])
        if not article:
            warnings.append("пустой артикул в позициях заказа")
            continue
        if quantity <= 0:
            warnings.append(f"количество для артикула {article} должно быть больше 0")
            continue
        items.append((article, quantity))

    return items, warnings


def parse_excel_date(value: Any, order_id: str) -> tuple[str | None, str | None]:
    if value is None or value == "":
        return None, None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d"), None
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d"), None

    if isinstance(value, (int, float)) and value > 0:
        # Excel хранит даты как число дней от 1899-12-30.
        parsed = datetime(1899, 12, 30) + timedelta(days=float(value))
        return parsed.strftime("%Y-%m-%d"), None

    text = clean_text(value)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d"), None
        except ValueError:
            continue

    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        return None, f"некорректная дата {text} заменена пустым значением"

    return text, None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value: Any) -> float:
    text = clean_text(value).replace(",", ".")
    if not text:
        return 0.0
    return float(text)


def to_int(value: Any) -> int:
    text = clean_text(value).replace(",", ".")
    if not text:
        return 0
    return int(float(text))


def write_import_report(path: Path, messages: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["Отчет импорта exam_db_master", ""]
    if messages:
        lines.extend(f"- {message}" for message in messages)
    else:
        lines.append("- Замечаний при импорте не найдено.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ============================================================
# РАЗДЕЛ: проверка БД и экспорт seed-данных
# ============================================================

def verify_database(settings: DbSettings) -> tuple[dict[str, int], list[str]]:
    issues: list[str] = []
    counts: dict[str, int] = {}

    with ExamDatabase(settings) as db:
        for table_name in TABLE_ORDER:
            try:
                row = db.execute(f"SELECT COUNT(*) AS count_value FROM {table_name}").fetchone()
                counts[table_name] = int(row["count_value"] if row else 0)
            except Exception as error:  # noqa: BLE001 - показываем понятную ошибку в verify.
                issues.append(f"Таблица {table_name}: {error}")

        if settings.engine == "sqlite":
            fk_rows = db.execute("PRAGMA foreign_key_check").fetchall()
            for row in fk_rows:
                issues.append(f"SQLite foreign_key_check: {row}")
        else:
            issues.extend(check_mysql_like_foreign_keys(db))

    required_non_empty = ("roles", "users", "products", "pickup_points", "orders", "order_items")
    for table_name in required_non_empty:
        if counts.get(table_name, 0) == 0:
            issues.append(f"Таблица {table_name} пустая.")

    return counts, issues


def check_mysql_like_foreign_keys(db: ExamDatabase) -> list[str]:
    checks = {
        "users.role_id": """
            SELECT COUNT(*) AS count_value
            FROM users LEFT JOIN roles ON roles.id = users.role_id
            WHERE roles.id IS NULL
        """,
        "products.unit_id": """
            SELECT COUNT(*) AS count_value
            FROM products LEFT JOIN units ON units.id = products.unit_id
            WHERE units.id IS NULL
        """,
        "products.supplier_id": """
            SELECT COUNT(*) AS count_value
            FROM products LEFT JOIN suppliers ON suppliers.id = products.supplier_id
            WHERE suppliers.id IS NULL
        """,
        "products.manufacturer_id": """
            SELECT COUNT(*) AS count_value
            FROM products LEFT JOIN manufacturers ON manufacturers.id = products.manufacturer_id
            WHERE manufacturers.id IS NULL
        """,
        "products.category_id": """
            SELECT COUNT(*) AS count_value
            FROM products LEFT JOIN categories ON categories.id = products.category_id
            WHERE categories.id IS NULL
        """,
        "orders.pickup_point_id": """
            SELECT COUNT(*) AS count_value
            FROM orders LEFT JOIN pickup_points ON pickup_points.id = orders.pickup_point_id
            WHERE pickup_points.id IS NULL
        """,
        "orders.status_id": """
            SELECT COUNT(*) AS count_value
            FROM orders LEFT JOIN order_statuses ON order_statuses.id = orders.status_id
            WHERE order_statuses.id IS NULL
        """,
        "orders.client_user_id": """
            SELECT COUNT(*) AS count_value
            FROM orders LEFT JOIN users ON users.id = orders.client_user_id
            WHERE orders.client_user_id IS NOT NULL AND users.id IS NULL
        """,
        "order_items.order_id": """
            SELECT COUNT(*) AS count_value
            FROM order_items LEFT JOIN orders ON orders.id = order_items.order_id
            WHERE orders.id IS NULL
        """,
        "order_items.product_id": """
            SELECT COUNT(*) AS count_value
            FROM order_items LEFT JOIN products ON products.id = order_items.product_id
            WHERE products.id IS NULL
        """,
    }

    issues = []
    for label, sql in checks.items():
        row = db.execute(sql).fetchone()
        count = int(row["count_value"] if row else 0)
        if count:
            issues.append(f"Нарушение связи {label}: {count}")
    return issues


def print_verify_report(counts: dict[str, int], issues: list[str]) -> None:
    counts_text = ", ".join(f"{table}={counts.get(table, 0)}" for table in TABLE_ORDER)
    if not issues:
        print(f"OK: {counts_text}")
        return

    print(f"Проверка завершена с проблемами: {counts_text}")
    for issue in issues:
        print(f"  - {issue}")


def export_seed(settings: DbSettings, output_path: Path) -> None:
    lines = [
        "-- Seed-файл создан exam_db_master.py",
        "SET FOREIGN_KEY_CHECKS = 0;" if settings.engine == "mysql" else "PRAGMA foreign_keys = OFF;",
        "",
    ]

    with ExamDatabase(settings) as db:
        for table_name in TABLE_ORDER:
            columns = TABLE_COLUMNS[table_name]
            rows = db.execute(
                f"SELECT {', '.join(columns)} FROM {table_name} ORDER BY id"
            ).fetchall()
            for row in rows:
                values = ", ".join(sql_literal(row[column]) for column in columns)
                lines.append(f"INSERT INTO {table_name}({', '.join(columns)}) VALUES ({values});")

    lines.extend(
        [
            "",
            "SET FOREIGN_KEY_CHECKS = 1;" if settings.engine == "mysql" else "PRAGMA foreign_keys = ON;",
        ]
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    if isinstance(value, datetime):
        value = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        value = value.strftime("%Y-%m-%d")
    text = str(value).replace("'", "''")
    return f"'{text}'"


# ============================================================
# РАЗДЕЛ: команды CLI
# ============================================================

def command_scan(args: argparse.Namespace) -> int:
    with open_source_workspace(Path(args.source)) as root:
        dataset = load_dataset(root)
        report = build_scan_report(dataset)
        print_scan_report(report)
    return 0


def command_init(args: argparse.Namespace) -> int:
    settings = db_settings_from_args(args)
    initialize_database(settings, reset=args.reset)
    print(f"OK: схема создана ({settings.engine}).")
    return 0


def command_import(args: argparse.Namespace) -> int:
    settings = db_settings_from_args(args)
    initialize_database(settings, reset=args.reset)

    report_messages: list[str] = []
    with open_source_workspace(Path(args.source)) as root:
        dataset = load_dataset(root)
        scan_report = build_scan_report(dataset)
        if scan_report.missing_files or scan_report.missing_columns:
            print_scan_report(scan_report)
            raise MasterError("Импорт остановлен: не хватает обязательных файлов или колонок.")

        with ExamDatabase(settings) as db:
            import_all_data(db, dataset, Path(args.assets_dir), report_messages)

    report_path = Path(args.report)
    write_import_report(report_path, report_messages)
    print(f"OK: импорт завершён, отчёт: {report_path}")
    counts, issues = verify_database(settings)
    print_verify_report(counts, issues)
    return 1 if issues else 0


def command_verify(args: argparse.Namespace) -> int:
    settings = db_settings_from_args(args)
    counts, issues = verify_database(settings)
    print_verify_report(counts, issues)
    return 1 if issues else 0


def command_export_seed(args: argparse.Namespace) -> int:
    settings = db_settings_from_args(args)
    export_seed(settings, Path(args.out))
    print(f"OK: seed сохранён в {args.out}")
    return 0


def command_wizard(_args: argparse.Namespace) -> int:
    print("Экзаменационный мастер БД")
    source = input("1) Папка или архив с Excel-ассетами: ").strip()
    engine = input("2) Движок БД [sqlite/mysql, Enter = sqlite]: ").strip().lower() or "sqlite"
    db_name = input("3) Файл SQLite или имя MySQL-БД [Enter = exam.db/exam_db]: ").strip()
    reset_answer = input("4) Пересоздать таблицы перед импортом? [y/N]: ").strip().lower()
    reset = reset_answer in {"y", "yes", "д", "да"}

    if not source:
        raise MasterError("Источник данных не указан.")

    settings = DbSettings(engine=engine)
    if engine == "sqlite":
        settings.db_path = Path(db_name or "exam.db")
    elif engine == "mysql":
        settings.database = db_name or os.environ.get("DB_NAME", "exam_db")
        settings.host = os.environ.get("DB_HOST", "localhost")
        settings.port = int(os.environ.get("DB_PORT", "3306"))
        settings.user = os.environ.get("DB_USER", "root")
        settings.password = os.environ.get("DB_PASSWORD", "")
    else:
        raise MasterError("Движок БД должен быть sqlite или mysql.")

    print("\nШаг scan")
    with open_source_workspace(Path(source)) as root:
        dataset = load_dataset(root)
        print_scan_report(build_scan_report(dataset))

        print("\nШаг init/import")
        initialize_database(settings, reset=reset)
        messages: list[str] = []
        with ExamDatabase(settings) as db:
            import_all_data(db, dataset, Path("app/assets"), messages)

    write_import_report(Path("import_report.txt"), messages)
    counts, issues = verify_database(settings)
    print_verify_report(counts, issues)
    return 1 if issues else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Микро-ПО для подготовки SQLite/MySQL-БД из экзаменационных Excel-ассетов."
    )
    parser.add_argument("--debug", action="store_true", help="Показать полный traceback при ошибке.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="Проверить папку или архив с исходными данными.")
    scan_parser.add_argument("--source", required=True, help="Папка, .zip или .rar с Excel-файлами.")
    scan_parser.set_defaults(func=command_scan)

    init_parser = subparsers.add_parser("init", help="Создать таблицы SQLite/MySQL.")
    add_db_arguments(init_parser)
    init_parser.add_argument("--reset", action="store_true", help="Перед созданием очистить существующую БД.")
    init_parser.set_defaults(func=command_init)

    import_parser = subparsers.add_parser("import", help="Создать схему и импортировать данные.")
    add_db_arguments(import_parser)
    import_parser.add_argument("--source", required=True, help="Папка, .zip или .rar с Excel-файлами.")
    import_parser.add_argument("--reset", action="store_true", help="Пересоздать таблицы перед импортом.")
    import_parser.add_argument(
        "--assets-dir",
        default="app/assets",
        help="Куда копировать фото; внутри будет создана папка products.",
    )
    import_parser.add_argument("--report", default="import_report.txt", help="Файл отчёта импорта.")
    import_parser.set_defaults(func=command_import)

    verify_parser = subparsers.add_parser("verify", help="Проверить таблицы, количества и связи.")
    add_db_arguments(verify_parser)
    verify_parser.set_defaults(func=command_verify)

    seed_parser = subparsers.add_parser("export-seed", help="Экспортировать INSERT-данные в SQL-файл.")
    add_db_arguments(seed_parser)
    seed_parser.add_argument("--out", required=True, help="Куда сохранить seed_sqlite.sql или seed_mysql.sql.")
    seed_parser.set_defaults(func=command_export_seed)

    wizard_parser = subparsers.add_parser("wizard", help="Интерактивный мастер из 4 вопросов.")
    wizard_parser.set_defaults(func=command_wizard)

    return parser


def add_db_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--engine",
        choices=("sqlite", "mysql"),
        default=os.environ.get("DB_ENGINE", "sqlite"),
        help="Движок БД: sqlite или mysql.",
    )
    parser.add_argument("--db", default=os.environ.get("DB_PATH", "exam.db"), help="Файл SQLite-БД.")
    parser.add_argument("--database", default=os.environ.get("DB_NAME", "exam_db"), help="Имя MySQL-БД.")
    parser.add_argument("--host", default=os.environ.get("DB_HOST", "localhost"), help="MySQL host.")
    parser.add_argument("--port", default=os.environ.get("DB_PORT", "3306"), help="MySQL port.")
    parser.add_argument("--user", default=os.environ.get("DB_USER", "root"), help="MySQL user.")
    parser.add_argument("--password", default=os.environ.get("DB_PASSWORD", ""), help="MySQL password.")


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        return int(args.func(args))
    except MasterError as error:
        print(f"Ошибка: {error}", file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        print("Остановлено пользователем.", file=sys.stderr)
        return 130
    except Exception as error:  # noqa: BLE001 - обычному пользователю нужен короткий текст.
        if getattr(args, "debug", False):
            raise
        print(f"Ошибка: {error}", file=sys.stderr)
        print("Для подробностей запустите с --debug.", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
